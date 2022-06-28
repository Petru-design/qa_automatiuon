import os

from openpyxl import load_workbook

from utils.comparators import recursive_attribute_compare, compare_texts


color_attrs = {
    "auto": bool,
    "idx_base": int,
    "index": str,
    "indexed": int,
    "namespace": type(None),
    "rgb": str,
    "tagname": str,
    "theme": int,
}

side_attrs = {
    "border_style": str,
    "color": str,
    "idx_base": int,
    "namespace": int,
    "style": str,
}

format_attributes = {
    "alignment": {
        "horizontal": type(None),
        "indent": float,
        "justifyLastLine": float,
        "readingOrder": float,
        "relativeIndent": float,
        "shrinkToFit": float,
        "textRotation": int,
        "vertical": int,
        "wrapText": type(None),
    },
    "border": {
        "bottom": side_attrs,
        "diagonal": side_attrs,
        "diagonalDown": bool,
        "diagonalUp": bool,
        "diagonal_direction": bool,
        "end": type(None),
        "horizontal": type(None),
        "left": side_attrs,
        "outline": bool,
        "right": side_attrs,
        "start": type(None),
        "top": side_attrs,
        "vertical": type(None),
    },
    "fill": {"bgColor": color_attrs, "fgColor": color_attrs, "patternType": str},
    "font": {
        "b": bool,
        "charset": type(None),
        "color": type(None),
        "condense": type(None),
        "extend": type(None),
        "family": type(None),
        "i": bool,
        "name": str,
        "outline": type(None),
        "scheme": type(None),
        "shadow": type(None),
        "strike": type(None),
        "sz": float,
        "u": type(None),
        "vertAlign": type(None),
    },
}


def generate_cells(workbook_1, workbook_2):
    for sheet_1, sheet_2 in zip(workbook_1.worksheets, workbook_2.worksheets):
        for column_1, column_2 in zip(sheet_1.columns, sheet_2.columns):
            for cell_1, cell_2 in zip(column_1, column_2):
                yield (cell_1, cell_2)


def test_xlsx_text(baseline_path, subject_path, result_path):
    baseline_wb = load_workbook(baseline_path)
    subject_wb = load_workbook(subject_path)

    for baseline_cell, subject_cell in generate_cells(baseline_wb, subject_wb):
        if baseline_cell.value == "" or subject_cell.value == "":
            assert baseline_cell.value == subject_cell.value
        else:
            compare_texts(
                baseline_cell.value,
                subject_cell.value,
                os.path.join(result_path, "text_result.txt"),
            )


def test_xlsx_format(baseline_path, subject_path, result_path):
    baseline_wb = load_workbook(baseline_path)
    subject_wb = load_workbook(subject_path)

    for baseline_cell, subject_cell in generate_cells(baseline_wb, subject_wb):
        recursive_attribute_compare(baseline_cell, subject_cell, format_attributes)
